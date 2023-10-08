import os
import sys
import re
import datetime
import subprocess
import errno
import time
import datetime
import shutil
from pathlib  import Path
from operator import itemgetter
from operator import attrgetter

import openpyxl
from openpyxl.styles import Alignment

import difflib


const_row_heigt = 12
const_col_width = 1


#/* 対象パス */
g_target_paths      = set([])

#/* 除外パス */
g_except_paths      = set([])


g_stop_on_copy      = 0
g_full_path         = 0
g_log_limit         = 0
g_revision1         = ""
g_revision2         = ""
g_path1             = ""
g_path2             = ""
g_path_logs         = []
g_repo_info         = None
g_out_path          = "out"
g_log_file_name     = ""
g_default_log       = 1
g_patch_mode        = 0

#/* かぞえチャオ関連 */
g_kazoe_path        = r"..\kazoeciao"
g_out_cas_file      = 1
g_kazoe_only        = 0
g_out_xlsx_file     = ""
g_kazoe_history     = None

re_log_line            = re.compile(r"^r([0-9]+)\s\|\s([^\|]+)\s\|\s([0-9]+)-([0-9]+)-([0-9]+)\s([0-9]+):([0-9]+):([0-9]+)\s[^\|]+\s\|\s([0-9]+)\sline")
re_change_line         = re.compile(r"^\s+([MADR])\s([^\s]+)")
re_change_file         = re.compile(r"\/([^\/]+)$")
re_kazoe_module        = re.compile(r"\"([^\"]+)\",\"([^\"]+)\",\"([^\"]+)\",([^,]+),([0-9]+),([0-9]+),([0-9]+),([0-9]+),([0-9]+)")
re_kazoe_module_after  = re.compile(r"\"([^\"]+)\", ,\"([^\"]+)\",([^,]+),([0-9]+),([0-9]+),([0-9]+),([0-9]+),([0-9]+)")
re_kazoe_single_module = re.compile(r"\"([^\"]+)\",\"([^\"]+)\",([^,]+),([0-9]+),([0-9]+),([0-9\.]+)")
re_kazoe_total         = re.compile(r"全ステップ数,\s*,\s*,\s*,([0-9]+),([0-9]+),([0-9]+),([0-9]+),([0-9]+)")
re_kazoe_single_total  = re.compile(r"全ステップ数,\s*,\s*,([0-9]+),([0-9]+),([0-9\.]+)")
re_cell_invalid_chr    = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")

#/*****************************************************************************/
#/* コミットログの中のファイル単位の情報の保持クラス                          */
#/*****************************************************************************/
class cChangeFile:
    def __init__(self):
        self.attribute = ""
        self.path      = ""
        self.file_name = ""
        self.external  = 0
        return


#/*****************************************************************************/
#/* svn のコミットログごとに得られる情報を保持するクラス                      */
#/*****************************************************************************/
class cCommitLog:
    def __init__(self):
        self.revision = 0
        self.author   = ""
        self.year     = 0
        self.month    = 0
        self.day      = 0
        self.hour     = 0
        self.minute   = 0
        self.second   = 0
        self.line     = 0
        self.refs     = []
        self.comments = []
        self.changes  = []
        return

    #/* コミットログの内容出力 */
    def output_log_text(self, target_path):
        file_path = target_path + "/commit_log.txt"
        with open(file_path, "w") as outfile:
            print("Revision : " + str(self.revision),  file = outfile)
            print("Author   : " + self.author,         file = outfile)
            print("Date     : %04d/%02d/%02d" % (self.year, self.month,  self.day),    file = outfile)
            print("Time     : %02d:%02d:%02d" % (self.hour, self.minute, self.second), file = outfile)
            print("", file = outfile)
            for comment in self.comments:
                print("Comment  : " + comment, file = outfile)

            print("", file = outfile)

            for change in self.changes:
                print("Files    : %s   %s" % (change.attribute, change.path), file = outfile)

            print("", file = outfile)
        return

    #/* コミットログの内容読み込み */
    def input_log_text(self, target_path):
        file_path = target_path + "/commit_log.txt"

        read_file = open(file_path, 'r')
        read_lines = read_file.readlines()

        for line in read_lines:
            if   (result := re.match(r"Revision : ([0-9]+)", line)):
                self.revision = int(result.group(1))
            elif (result := re.match(r"Author   : ([^\n]+)", line)):
                self.author   = result.group(1)
            elif (result := re.match(r"Date     : ([0-9]+)\/([0-9]+)\/([0-9]+)", line)):
                self.year     = int(result.group(1))
                self.month    = int(result.group(2))
                self.day      = int(result.group(3))
            elif (result := re.match(r"Time     : ([0-9]+)\/([0-9]+)\/([0-9]+)", line)):
                self.hour     = int(result.group(1))
                self.minute   = int(result.group(2))
                self.second   = int(result.group(3))
            elif (result := re.match(r"Comment  : ([^\n]+)", line)):
                self.comments.append(result.group(1))
        return


#/*****************************************************************************/
#/* svn logで得られる情報の保持クラス                                         */
#/*****************************************************************************/
class cPathLog:
    def __init__(self):
        self.path      = ""
        self.option    = ""
        self.logs      = []
        self.targets   = []
        self.first_rev = 0
        self.last_rev  = 0
        return

    def sort_revision(self):
        self.logs.sort(key=attrgetter('revision'))
        if not self.logs:
            self.first_rev = 0
            self.last_rev  = 0
        else:
            self.first_rev = self.logs[0].revision
            self.last_rev  = self.logs[-1].revision

        return


#/*****************************************************************************/
#/* svn infoで得られる情報の保持クラス                                        */
#/*****************************************************************************/
class cRepoInfo:
    def __init__(self):
        self.path         = ""
        self.url          = ""
        self.relative     = ""
        self.relative_dir = ""
        self.root         = ""
        self.node_kind    = ""
        return


#/*****************************************************************************/
#/* かぞえチャオのステップ数情報保持クラス                                    */
#/*****************************************************************************/
class cKazoeSteps:
    def __init__(self):
        self.real_steps   = 0                  #/* 実ステップ数          */
        self.new_steps    = 0                  #/* 新規ステップ          */
        self.base_steps   = 0                  #/* 流用元ステップ        */
        self.mod_steps    = 0                  #/* 変更ステップ          */
        self.div_steps    = 0                  #/* 流用ステップ          */
        self.del_steps    = 0                  #/* 削除ステップ          */
        return

    #/* 差分ステップの登録 */
    def set_diff_steps(self, new_steps, base_steps, mod_steps, div_steps, del_steps):
        self.new_steps    = new_steps
        self.base_steps   = base_steps
        self.mod_steps    = mod_steps
        self.div_steps    = div_steps
        self.del_steps    = del_steps
        self.real_steps   = new_steps + mod_steps + div_steps    #http://ciao-ware.c.ooco.jp/ft_faq_kazo004.html
        return

    #/* ステップの加算登録 */
    def add_steps(self, steps):
        self.real_steps   += steps.real_steps
        self.new_steps    += steps.new_steps
        self.base_steps   += steps.base_steps
        self.mod_steps    += steps.mod_steps
        self.div_steps    += steps.div_steps
        self.del_steps    += steps.del_steps
        return


#/*****************************************************************************/
#/* かぞえチャオの関数（モジュール） 情報保持クラス                           */
#/*****************************************************************************/
class cKazoeModule:
    def __init__(self):
        self.before_path  = ""
        self.after_path   = ""
        self.file_name    = ""
        self.module_name  = ""
        self.module_type  = ""
        self.steps        = None
        return


#/*****************************************************************************/
#/* かぞえチャオの結果をソースファイルごとに集計した情報を保持するクラス      */
#/*****************************************************************************/
class cKazoeFile:
    def __init__(self):
        self.file_name    = ""
        self.file_steps   = cKazoeSteps()
        self.modules      = []
        return

    #/* モジュールステップ行の処理 */
    def add_one_module_line(self, after_path, before_path, module_name, module_type, steps):
        module = cKazoeModule()
        module.after_path       = after_path
        module.before_path      = before_path
        module.module_name      = module_name
        module.module_type      = module_type
        module.steps            = steps
        if (before_path != ""):
            print("%s, %s, %s, %d, %d, %d, %d, %d, %d" % (module.before_path, module.after_path, module.module_name, module.steps.real_steps, module.steps.new_steps, module.steps.base_steps, module.steps.mod_steps, module.steps.div_steps, module.steps.del_steps))
        else:
            print("%s, %s, %d" % (module.after_path, module.module_name, module.steps.real_steps))
        self.modules.append(module)
        self.file_steps.add_steps(steps)
        return


#/*****************************************************************************/
#/* かぞえチャオの結果ファイル（csv）情報保持クラス                           */
#/*****************************************************************************/
class cKazoeResult:
    def __init__(self):
        self.rslt_file    = ""                 #/* CSVファイルの相対パス */
        self.before_rev   = 0                  #/* 変更前のRevision      */
        self.after_rev    = 0                  #/* 変更後のRevision      */
        self.modules      = []
        self.files        = []
        self.total_steps  = None               #/* TotalのStep数情報     */
        self.commit_log   = None               #/* cCommitLogへの参照    */
        return

    #/* ファイル名からcKazoeFileクラスのインスタンスを取得 */
    def find_file_class(self, file_name):
        for file in self.files:
            if (file_name == file.file_name):
                return file

        new_file = cKazoeFile()
        new_file.file_name = file_name
        self.files.append(new_file)
        return new_file

    #/* モジュールステップ行の処理 */
    def add_one_module_line(self, after_path, before_path, module_name, module_type, steps):
        rslt_path = os.path.dirname(self.rslt_file).replace("\\", "\\\\") + r"\\export"
        file_name = ""

        #/* 絶対パスからファイル名を取得する */
        result = re.search(rslt_path + r"\\(.+)$", after_path)
        if (result):
            file_name = result.group(1)
            file = self.find_file_class(file_name)
            file.add_one_module_line(after_path, before_path, module_name, module_type, steps)
        else:
            print("no hit %s : %s" % (after_path, rslt_path + r"\\(.+)$"))
        return


#/*****************************************************************************/
#/* 履歴情報出力クラス                                                        */
#/*****************************************************************************/
class cKazoeHistory:
    def __init__(self):
        self.first_rev    = 0                  #/* 最初のリビジョン      */
        self.last_rev     = 0                  #/* 最後のリビジョン      */
        self.url          = 0                  #/* リポジトリURL         */
        self.rslts        = []
        return

    def sort_revision(self):
        self.rslts.sort(key=attrgetter('after_rev'))

        if not self.rslts:
            self.first_rev = 0
            self.last_rev  = 0
        else:
            self.first_rev = self.rslts[0].after_rev
            self.last_rev  = self.rslts[-1].after_rev
        return


#/*****************************************************************************/
#/* ログファイル設定                                                          */
#/*****************************************************************************/
def log_settings():
    global g_log_file_name
    global g_out_path
    global g_default_log
    global g_out_xlsx_file


    make_directory(g_out_path)
    log_path = ""

#   print("log_settings")
    if (g_log_file_name != ""):
        log_path = g_out_path + "\\" + g_log_file_name
    elif (g_default_log == 1):
        now = datetime.datetime.now()
        formatted_time = now.strftime("%Y%m%d_%H%M%S")
        log_path = g_out_path + "\\svn_checker_" + formatted_time + ".log";
        g_out_xlsx_file = g_out_path + "\\svn_checker_" + formatted_time + ".xslx";

    print ("log_path : %s" % log_path)

    if (log_path != ""):
       log_file = open(log_path, "a")
       sys.stdout = log_file

    now = datetime.datetime.now()
    print("start checking : " + str(now))

    return

#/*****************************************************************************/
#/* 外部コマンド実行                                                          */
#/*****************************************************************************/
def cmd_execute(cmd_text, out_file, in_file):
    if (out_file == ""):
        if (in_file == ""):
            result = subprocess.run(cmd_text, capture_output=True, text=True)
        else:
            with open(in_file, "r") as infile:
                result = subprocess.run(cmd_text, capture_output=True, text=True, stdin=infile)
    else:
        if (in_file == ""):
            with open(out_file, "w") as outfile:
                result = subprocess.run(cmd_text, stdout=outfile)
        else:
            with open(out_file, "w") as outfile:
                with open(in_file, "r") as infile:
                    result = subprocess.run(cmd_text, stdout=outfile, stdin=infile)

#   print(result.stdout)
    return result.stdout


#/*****************************************************************************/
#/* ディレクトリ作成                                                          */
#/*****************************************************************************/
def make_directory(path):
    try:
        os.makedirs(path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise
    return

#/*****************************************************************************/
#/* ディレクトリ作成                                                          */
#/*****************************************************************************/
def force_copy_directory(src_path, dst_path):
    if os.path.exists(dst_path):
        shutil.rmtree(dst_path)

    try:
        shutil.copytree(src_path, dst_path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise
    return


#/*****************************************************************************/
#/* ターゲットパス判定                                                        */
#/*****************************************************************************/
def is_path_in_target(path):
    global g_repo_info
    global g_target_paths

    #/* 除外パスにヒットしたらFalse */
    for except_path in g_except_paths:
        if (re.search(except_path, path)):
            return False

    #/* relative_dirそのものはターゲットから除外する */
    if (path == g_repo_info.relative_dir):
        return False

    #/* ターゲットパス指定が空の場合は、trueを返す */
    if not g_target_paths:
        return True

    #/* ターゲットパスにヒットしたらTrue */
    for target_path in g_target_paths:
        if (re.search(target_path, path)):
            return True

    return False


#/*****************************************************************************/
#/* コマンドライン引数処理                                                    */
#/*****************************************************************************/
def check_command_line_option():
    global g_stop_on_copy
    global g_log_limit
    global g_kazoe_path
    global g_revision1
    global g_revision2
    global g_path1
    global g_path2
    global g_out_path
    global g_full_path
    global g_kazoe_only
    global g_out_cas_file
    global g_kazoe_history
    global g_target_paths
    global g_except_paths

    argc = len(sys.argv)
    option = ""

    if (argc == 1):
        print("usage ; svn_checker.py")
        exit(-1)

    sys.argv.pop(0)
    for arg in sys.argv:
        if (option == "r"):
            if (g_revision1 == ""):
                g_revision1 = arg
            elif (g_revision2 == ""):
                g_revision2 = arg
            else:
                print("svn_checker.py : Too many revisions!")
                exit(-1)
            option = ""
        elif (option == "l"):
            g_log_limit = int(arg)
#           print("log limit %s" % g_log_limit)
            option = ""
        elif (option == "o"):
            g_out_path = arg
            option = ""
        elif (option == "k"):
            g_kazoe_path = arg
            option = ""
        elif (option == "t"):
            g_target_paths.add(arg)
            option = ""
        elif (option == "e"):
            g_except_paths.add(arg)
            option = ""
        elif (arg == "--stop-on-copy"):
            g_stop_on_copy = 1
        elif (arg == "-f") or (arg == "--fullpath"):
            g_full_path = 1
        elif (arg == "-l") or (arg == "--limit"):
            option = "l"
        elif (arg == "-r") or (arg == "--revision"):
            option = "r"
        elif (arg == "-k") or (arg == "--kazoe"):
            option = "k"
        elif (arg == "-o"):
            option = "o"
        elif (arg == "-e"):
            option = "e"
        elif (arg == "-t"):
            option = "t"
        elif (arg == "-ko"):
            g_kazoe_only = 1
        elif (g_path1 == ""):
            if (result := re.search(r"\/$", arg)):
#               print("end by /")
                g_path1 = arg[:-1]
            else:
                g_path1 = arg
        elif (g_path2 == ""):
            g_path2 = arg
        else:
            print("svn_checker.py : Too many paths!")
            exit(-1)


    if (g_path1 == ""):
        print("svn_checker.py : no target path!")
        exit(-1)

    if (g_out_cas_file != 0):
        g_kazoe_history = cKazoeHistory()
        g_kazoe_history.url = g_path1

    return



#/*****************************************************************************/
#/* svnログ取得                                                               */
#/*****************************************************************************/
def check_log(target_path, revision, limit):
    global g_path_logs
    global g_repo_info

    cmd_text = r"svn log -v "
    option   = ""

    if (revision != ""):
        option += r"-r " + revision + " "

    if (limit != 0):
        option += r"-l " + str(limit) + " "

    cmd_text += option + target_path
    path_log        = cPathLog()
    path_log.path   = target_path
    path_log.option = option

    print(cmd_text)
    lines = cmd_execute(cmd_text, "", "").split("\n")

    commit_log = None
    log_sequence = 0
    log_lines = 0
    for line in lines:
#       print("line : " + line)
        if (log_sequence == 0) or (log_sequence == 1):
            if (line == "------------------------------------------------------------------------"):
                log_sequence = 1
                if (commit_log != None):
                    path_log.logs.append(commit_log)
                    commit_log = None
            elif (result := re_log_line.match(line)):
                if (commit_log != None):
                    path_log.logs.append(commit_log)
                    commit_log = None

                commit_log = cCommitLog()
                commit_log.revision = int(result.group(1))
                commit_log.author   = result.group(2)
                commit_log.year     = int(result.group(3))
                commit_log.month    = int(result.group(4))
                commit_log.day      = int(result.group(5))
                commit_log.hour     = int(result.group(6))
                commit_log.minute   = int(result.group(7))
                commit_log.second   = int(result.group(8))
                commit_log.line     = int(result.group(9))
                print("rev : " + result.group(1) + ", author : " + result.group(2) + ", line : " + result.group(9))
#               print("datetime : %04d/%02d/%02d %02d:%02d:%02d" % (commit_log.year, commit_log.month, commit_log.day, commit_log.hour, commit_log.minute, commit_log.second))
                log_sequence = 2
                log_lines    = commit_log.line
            elif (line == ""):
                pass
            else:
                print("strange log in [%d] : %s" % (log_sequence, line))
        elif (log_sequence == 2):
            if (line == "Changed paths:"):
                log_sequence = 3
            else:
                print("strange log in [%d] : %s" % (log_sequence, line))
        elif (log_sequence == 3):
            if (line == ""):
                log_sequence = 4
            elif (result := re_change_line.match(line)):
                changed = cChangeFile()
                changed.attribute = result.group(1)
                changed.path      = result.group(2)
                
                if (result := re_change_file.search(changed.path)):
                    changed.file_name = result.group(1)
                else:
                    changed.file_name = ""
                if (re.match(g_repo_info.relative, changed.path) ):
                    #/* 指定されたリポジトリパスの範囲内のファイル */
                    print("in type : %s, file : %s, path : %s" % (changed.attribute, changed.file_name, changed.path))
                    changed.external = 0
                    path_log.targets.append(changed.path)
                else:
                    #/* 指定されたリポジトリパスの範囲外のファイル */
#                   print("ex type : %s, file : %s, path : %s" % (changed.attribute, changed.file_name, changed.path))
                    changed.external = 1
                commit_log.changes.append(changed)
            else:
                print("strange log in [%d] : %s" % (log_sequence, line))
        elif (log_sequence == 4):
            commit_log.comments.append(line)
            log_lines -= 1
            if (log_lines == 0):
                log_sequence = 0
            pass

    #/* 最後にターゲットpathの重複を削除して、ログをRevisionの昇順でソートして登録する */
    path_log.targets = set(path_log.targets)
    path_log.sort_revision()
    g_path_logs.append(path_log)
    return





#/*****************************************************************************/
#/* パスのログ確認                                                            */
#/*****************************************************************************/
def check_path_log():
    global g_path1
    global g_path2
    global g_log_limit
    global g_revision1
    global g_revision2

    if (g_path2 == ""):
        check_log(g_path1, g_revision1, g_log_limit)
    else:
        check_log(g_path1)

    return


#/*****************************************************************************/
#/* リポジトリ情報の確認                                                      */
#/*****************************************************************************/
def check_repo_info():
    global g_path1
    global g_repo_info

    g_repo_info = cRepoInfo()
    lines = cmd_execute("svn info " + g_path1, "", "").split("\n")
    for line in lines:
#       print(line)
        if (result := re.match(r"Path: ([^\s]+)", line)):
            g_repo_info.path = result.group(1)
        elif (result := re.match(r"URL: ([^\s]+)", line)):
            g_repo_info.url = result.group(1)
        elif (result := re.match(r"Relative URL: \^([^\s]+)", line)):
            g_repo_info.relative = result.group(1)
        elif (result := re.match(r"Repository Root: ([^\s]+)", line)):
            g_repo_info.root = result.group(1)
        elif (result := re.match(r"Node Kind: ([^\s]+)", line)):
            g_repo_info.node_kind = result.group(1)

    if (g_path1 != g_repo_info.url):
        print("strange svn info. URL unmatch! [%s] - [%s]" % (g_path1, g_repo_info.url))
        exit(-1)

    print("Path : %s, URL : %s, Root : %s, Relative : %s, Kind : %s" % (g_repo_info.path, g_repo_info.url, g_repo_info.root, g_repo_info.relative, g_repo_info.node_kind))
    if (g_repo_info.node_kind == "file"):
        g_repo_info.relative_dir = os.path.dirname(g_repo_info.relative)
        print("  Relative Directory : %s" % (g_repo_info.relative_dir))
    else:
        g_repo_info.relative_dir = g_repo_info.relative
    return


#/*****************************************************************************/
#/* かぞえチャオ出力ファイルの削除                                            */
#/*****************************************************************************/
def clean_ciao_rslt(target_path):
    files = os.listdir(target_path)

    for filename in files:
        if (os.path.isfile(target_path + "\\" + filename)):
            if (result := re.match(r"^ciao_rslt[0-9]+_[0-9]+.csv",filename)):
                print("remove ciao rslt file : %s" % (target_path + "\\" + filename))
                os.remove(target_path + "\\" + filename)

    return


#/*****************************************************************************/
#/* かぞえチャオ出力ファイルの確認                                            */
#/*****************************************************************************/
def find_ciao_rslt(target_path):
    files = os.listdir(target_path)

    for filename in files:
        if (os.path.isfile(target_path + "\\" + filename)):
            if (result := re.match(r"^ciao_rslt[0-9]+_[0-9]+.csv",filename)):
                print("find ciao rslt file : %s" % (target_path + "\\" + filename))
                return (target_path + "\\" + filename)

            if (result := re.match(r"^ciao_rslt_r[0-9]+_r[0-9]+.csv",filename)):
                print("find ciao rslt file : %s" % (target_path + "\\" + filename))
                return (target_path + "\\" + filename)

    return ""


#/*****************************************************************************/
#/* かぞえチャオ出力ファイルの確認(outpath全検索)                             */
#/*****************************************************************************/
def find_ciao_rslt_all():
    global g_kazoe_history
    before_rev = 0

    #/* かぞえチャオ結果確認のみの場合、ここで結果のcsvファイルをチェックしていく */
    if (g_kazoe_only == 1):
        files = os.listdir(g_out_path)

        for dir_name in files:
            if (os.path.isdir(g_out_path + "\\" + dir_name)):
                if (result := re.match(r"^rev_([0-9]+)",dir_name)):
                    rslt_file = find_ciao_rslt(g_out_path + "\\" + dir_name)
                    if (rslt_file != ""):
                        rslt = cKazoeResult()
                        rslt.rslt_file = rslt_file
                        rslt.after_rev = int(result.group(1))

                        commit_log = cCommitLog()
                        commit_log.input_log_text(g_out_path + "\\" + dir_name)
                        rslt.commit_log = commit_log
                        g_kazoe_history.rslts.append(rslt)
                    else:
                        before_rev = result.group(1)


    #/* 変更後Revisionでソートする */
#   g_kazoe_history.rslts.sort(key=attrgetter('after_rev'))
    g_kazoe_history.sort_revision()


    #/* かぞえチャオ結果確認のみの場合、さらに変更前Revisionをセットする */
    if (g_kazoe_only == 1):
        for rslt in g_kazoe_history.rslts:
            rslt.before_rev = before_rev
            before_rev = rslt.after_rev

    return


#/*****************************************************************************/
#/* かぞえチャオ出力結果の読み込み                                            */
#/*****************************************************************************/
def read_ciao_rslt(rslt):
    csv = open(rslt.rslt_file, 'r')
    read_lines = csv.readlines()
    rslt_path = os.path.dirname(rslt.rslt_file).replace("\\", "\\\\") + r"\\export"
    print("read kazoe result file for rev %d : %s : %s" % (rslt.after_rev, rslt.rslt_file, rslt_path))

    read_lines.pop(0)
    for line in read_lines:
#       print(line)
        if (result := re_kazoe_module.match(line)):
            steps = cKazoeSteps()
            steps.set_diff_steps(int(result.group(5)), int(result.group(6)), int(result.group(7)), int(result.group(8)), int(result.group(9)))
            rslt.add_one_module_line(result.group(1), result.group(2), result.group(3), result.group(4), steps)
        elif (result := re_kazoe_module_after.match(line)):
            steps = cKazoeSteps()
            steps.set_diff_steps(int(result.group(4)), int(result.group(5)), int(result.group(6)), int(result.group(7)), int(result.group(8)))
            rslt.add_one_module_line(result.group(1), " ", result.group(2), result.group(3), steps)
        elif (result := re_kazoe_single_module.match(line)):
            steps = cKazoeSteps()
            steps.real_steps = int(result.group(5))
            rslt.add_one_module_line(result.group(1), "", result.group(2), result.group(3), steps)
        elif (result := re_kazoe_total.match(line)):
            steps = cKazoeSteps()
            steps.set_diff_steps(int(result.group(1)), int(result.group(2)), int(result.group(3)), int(result.group(4)), int(result.group(5)))
            rslt.total_steps = steps
            print("total steps : %d, %d, %d, %d, %d" % (rslt.total_steps.new_steps, rslt.total_steps.base_steps, rslt.total_steps.mod_steps, rslt.total_steps.div_steps, rslt.total_steps.del_steps))
        elif (result := re_kazoe_single_total.match(line)):
            steps = cKazoeSteps()
            steps.real_steps = int(result.group(2))
            rslt.total_steps = steps
            print("total steps : %d" % (rslt.total_steps.real_steps))

    csv.close()
    return


#/*****************************************************************************/
#/* かぞえチャオ自動実行ファイル（*.cas）作成と実行                           */
#/*****************************************************************************/
def output_cas_text(commit_log, pre_revision, target_path, before, after):
    global g_out_cas_file
    global g_kazoe_history

    revision = commit_log.revision
    if (g_out_cas_file != 0):
        target_path = os.getcwd() + "\\" + target_path.replace("/", "\\")
        file_path = target_path + "/diff_kazoe.cas"
        with open(file_path, "w") as outfile:
                after_path  = os.getcwd() + "\\" + after.replace("/", "\\")
                before_path = os.getcwd() + "\\" + before.replace("/", "\\")
                print(r"[AFTPATH]", file = outfile)
                print(after_path,   file = outfile)
                print(r"[BFRPATH]", file = outfile)
                print(before_path,  file = outfile)
                print(r"[EXEMODE]", file = outfile)
                if (pre_revision != 0):
                    print(r"1",  file = outfile)
                else:
                    print(r"2",  file = outfile)
                print(r"[RSLMODE]", file = outfile)
                print(r"0",  file = outfile)
                print(r"[RSLPATH]", file = outfile)
                print(target_path,  file = outfile)
                outfile.close()
                clean_ciao_rslt(target_path)

                kazoe_cmd = g_kazoe_path + "\\" + "kazoeciao.exe /a" + file_path
                print(kazoe_cmd)
                lines = cmd_execute(kazoe_cmd, "", "")

                rslt_file = find_ciao_rslt(target_path)
                rename_file = os.path.dirname(rslt_file) + "\\ciao_rslt_r" + str(pre_revision) + "_r" + str(revision) + ".csv"
                if (rslt_file != ""):
                    print("rename [%s] to [%s]" % (rslt_file, rename_file))
                    os.rename(rslt_file, rename_file)
                    rslt = cKazoeResult()
                    rslt.rslt_file  = rename_file
                    rslt.after_rev  = revision
                    rslt.before_rev = pre_revision
                    rslt.commit_log = commit_log
                    g_kazoe_history.rslts.append(rslt)

    return


#/*****************************************************************************/
#/* ログ内のターゲットファイル出力                                            */
#/*****************************************************************************/
def output_log_files(path_log, commit_log, pre_revision):
    global g_repo_info
    global g_full_path
    global g_out_path
    global g_patch_mode

    rev_path = g_out_path + "/rev_" + str(commit_log.revision)
    print("output_log_files for rev:%d, pre_rev: %d" % (commit_log.revision, pre_revision))

    export_path = rev_path + "/export"
    out_path = rev_path
    pre_path = g_out_path + "/rev_" + str(pre_revision)

    if ((pre_revision == 0) or (g_patch_mode == 0)):
        #/* 初回のRevisionもしくはパッチモードOFFの場合は全ファイルexportで取得 */
        cmd_base = "svn export -r " + str(commit_log.revision) + " "
        make_directory(export_path)
        for target in path_log.targets:
            if (is_path_in_target(target)):
                print("out for %d : %s" % (commit_log.revision, target))
                if (g_full_path):
                    out_path = export_path + os.path.dirname(target)
                    export_cmd = cmd_base + g_repo_info.root + target + " " + out_path
                else:
                    out_path = export_path + os.path.dirname(target).replace(g_repo_info.relative_dir, "")
                    export_cmd = cmd_base + g_repo_info.root + target + " " + out_path
                print("create path : %s, relative : %s" % (out_path, g_repo_info.relative_dir))
                make_directory(out_path)
                print("export : %s" % (export_cmd))
                lines = cmd_execute(export_cmd, "", "")
    else:
        #/* 2回目以降のRevisionに対しては、svn diffによるPatchを取得して、Patchを当てていくことで高速化する */
        force_copy_directory(pre_path, rev_path)
        out_diff = out_path + "/diff_from_r" + str(pre_revision) + ".diff"
        diff_cmd = "svn diff -r " + str(pre_revision) + ":" + str(commit_log.revision) + " " + g_repo_info.url
        print(diff_cmd)
        lines = cmd_execute(diff_cmd, out_diff, "")
        print(lines)
        if (g_full_path):
            patch_cmd = "patch -d " + export_path + g_repo_info.relative_dir + " -p0"
        else:
            patch_cmd = "patch -d " + export_path + " -p0"
        print(patch_cmd)
        lines = cmd_execute(patch_cmd, "", out_diff)
        os.remove(out_diff)

    #/* かぞえチャオ自動実行 */
    output_cas_text(commit_log, pre_revision, rev_path, pre_path + "/export", export_path)

    #/* コミットログの内容をテキストファイルに出力 */
    commit_log.output_log_text(rev_path)
    return


#/*****************************************************************************/
#/* パス内のファイル履歴出力                                                  */
#/*****************************************************************************/
def output_path_files():
    global g_repo_info
    global g_path_logs
    global g_out_path
    global g_full_path

    pre_revision = 0
    make_directory(g_out_path)
    for path_log in g_path_logs:
        for target in path_log.targets:
            print("target : %s" % target)

        for log in path_log.logs:
#           print("log for rev : %d" % log.revision)
            for changed in log.changes:
                if (changed.external == 0):
#                   print("hit on rev %d" % log.revision)
                    if (is_path_in_target(changed.path)):
                        output_log_files(path_log, log, pre_revision)
                        pre_revision = log.revision
                        break

    return


#/*****************************************************************************/
#/* かぞえチャオの出力ファイルの読み出し                                      */
#/*****************************************************************************/
def check_kazoe_result():
    global g_kazoe_history
    global g_out_cas_file

    if (g_out_cas_file != 0):
        find_ciao_rslt_all()

        for rslt in g_kazoe_history.rslts:
            read_ciao_rslt(rslt)

    return


#/*****************************************************************************/
#/* シートの存在確認                                                          */
#/*****************************************************************************/
def check_ws_exists(wb, sheet_name):
    for ws in wb.worksheets:
        if (ws.title == sheet_name):
            return True

    return False


#/*****************************************************************************/
#/* セルのテキスト検索                                                        */
#/*****************************************************************************/
def find_row(ws, col, row, name):
    while (ws.cell(row, col).value != None):
        if (ws.cell(row, col).value == name):
            return row

        row += 1

    ws.cell(row, col).value = name
    return row


#/*****************************************************************************/
#/* かぞえチャオの情報履歴の出力                                              */
#/*****************************************************************************/
def out_kazoe_history():
    global g_kazoe_history
    global g_out_cas_file
    global g_out_xlsx_file

    if (g_out_cas_file != 0):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = "ファイル履歴"

        col = 2
        ws.cell(2, col).value = g_kazoe_history.url
        ws.cell(3, col).value = r"Rev." + str(g_kazoe_history.first_rev) + r" - " + r"Rev." + str(g_kazoe_history.last_rev)
        ws.cell(5, col).value = "Revision"
        ws.cell(6, col).value = "Author"
        ws.cell(7, col).value = "Comment"
        ws.cell(8, col).value = "全体"
        ws.cell(9, col).value = "変更ファイル"

        rev_count         = 0
        max_comment_count = 1
        max_line_count    = 1
        max_file_name     = 9
        file_with_modules = set()

        #/* ファイル単位のステップカウント履歴を作成 */
        for rslt in g_kazoe_history.rslts:

            #/* セル結合してリビジョンを出力 */
            ws.merge_cells(start_row= 5, end_row=5, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
            ws.cell(5, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left")
            ws.cell(5, col + (rev_count * 3) + 1).value = rslt.commit_log.revision

            #/* セル結合してAuthorを出力 */
            ws.merge_cells(start_row= 6, end_row=6, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
            ws.cell(6, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left")
            ws.cell(6, col + (rev_count * 3) + 1).value = rslt.commit_log.author

            #/* セル結合してコメントを出力 */
            ws.merge_cells(start_row= 7, end_row=7, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
            ws.cell(7, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left", vertical="top")
            comment = ""
            comment_count = 1
            for comment_line in rslt.commit_log.comments:
                if (comment == ""):
                    comment = comment_line
                else:
                    comment += "\n" + comment_line
                    comment_count += 1
            try:
                ws.cell(7, col + (rev_count * 3) + 1).value = comment
            except openpyxl.utils.exceptions.IllegalCharacterError as exception:
                ws.cell(7, col + (rev_count * 3) + 1).value = "例外発生"
            if (comment_count > max_comment_count):
                ws.row_dimensions[7].height = comment_count * const_row_heigt
                max_comment_count = comment_count

            #/* 全体のステップ数（新規＋変更、削除、実ステップ）を出力 */
            ws.cell(8, col + (rev_count * 3) + 1).value = rslt.total_steps.new_steps + rslt.total_steps.mod_steps
            ws.cell(8, col + (rev_count * 3) + 2).value = rslt.total_steps.del_steps
            ws.cell(8, col + (rev_count * 3) + 3).value = rslt.total_steps.real_steps

            #/* ファイル単位のステップ数出力 */
            changed_files = []
            for file_rslt in rslt.files:
                row = find_row(ws, col, 10, file_rslt.file_name)
                ws.cell(row, col + (rev_count * 3) + 1).value = file_rslt.file_steps.new_steps + file_rslt.file_steps.mod_steps
                ws.cell(row, col + (rev_count * 3) + 2).value = file_rslt.file_steps.del_steps
                ws.cell(row, col + (rev_count * 3) + 3).value = file_rslt.file_steps.real_steps

                #/* 最長ファイル名の長さに合わせてB列幅を調整 */
                if (len(file_rslt.file_name) > max_file_name):
                    max_file_name = len(file_rslt.file_name)
                    ws.column_dimensions['B'].width = max_file_name * const_col_width

                #/* 変更のあったファイル名をピックアップ */
                if ((file_rslt.file_steps.new_steps + file_rslt.file_steps.mod_steps + file_rslt.file_steps.del_steps) > 0):
                    changed_file = os.path.basename(file_rslt.file_name)
                    changed_files.append(changed_file)

                #/* 複数モジュールを持つファイルをピックアップしておく */
                if (len(file_rslt.modules) > 1):
                    length_before = len(file_with_modules)
                    file_with_modules.add(file_rslt.file_name)
                    if (len(file_with_modules) > length_before):
                        print("add file with modules [%s] in Rev.%d" % (file_rslt.file_name, rslt.commit_log.revision))

            #/* セル結合して変更のあったモジュールの出力 */
            ws.merge_cells(start_row= 9, end_row=9, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
            ws.cell(9, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left", vertical="top")
            changed = ""
            line_count = 1
            for changed_file in changed_files:
                if (changed == ""):
                    changed = changed_file
                else:
                    changed += "\n" + changed_file
                    line_count += 1
            ws.cell(9, col + (rev_count * 3) + 1).value = changed

            if (line_count > max_line_count):
                ws.row_dimensions[9].height = line_count * const_row_heigt
                max_line_count = line_count

            rev_count += 1


        #/* ファイル毎にモジュール単位のステップカウント履歴を作成 */
        for file_with_module in file_with_modules:
            file_name = os.path.basename(file_with_module)
            if (len(file_name) > 31):
                print("too long file name! [%s]" % (file_name))
                file_name = file_name[:31]

            letter = 0
            while (check_ws_exists(wb, file_name)):
                print("Sheet Name Duplication! [%s]" % (file_name))
                file_name = file_name[:30]
                file_name += chr(ord('A') + letter)
                letter += 1

            ws = wb.create_sheet(file_name)

            col = 2
            ws.cell(2, col).value = file_with_module
            ws.cell(3, col).value = r"Rev." + str(g_kazoe_history.first_rev) + r" - " + r"Rev." + str(g_kazoe_history.last_rev)
            ws.cell(5, col).value = "Revision"
            ws.cell(6, col).value = "Author"
            ws.cell(7, col).value = "Comment"
            ws.cell(8, col).value = "全体"
            ws.cell(9, col).value = "変更関数"

            rev_count         = 0
            max_comment_count = 1
            max_line_count    = 1
            max_file_name     = 9

            #/* ファイル単位のステップカウント履歴を作成 */
            for rslt in g_kazoe_history.rslts:

                #/* セル結合してリビジョンを出力 */
                ws.merge_cells(start_row= 5, end_row=5, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
                ws.cell(5, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left")
                ws.cell(5, col + (rev_count * 3) + 1).value = rslt.commit_log.revision

                #/* セル結合してAuthorを出力 */
                ws.merge_cells(start_row= 6, end_row=6, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
                ws.cell(6, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left")
                ws.cell(6, col + (rev_count * 3) + 1).value = rslt.commit_log.author

                #/* セル結合してコメントを出力 */
                ws.merge_cells(start_row= 7, end_row=7, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
                ws.cell(7, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left", vertical="top")
                comment = ""
                comment_count = 1
                for comment_line in rslt.commit_log.comments:
                    if (comment == ""):
                        comment = comment_line
                    else:
                        comment += "\n" + comment_line
                        comment_count += 1
                re_cell_invalid_chr.sub(" ", comment)
                try:
                    ws.cell(7, col + (rev_count * 3) + 1).value = comment
                except openpyxl.utils.exceptions.IllegalCharacterError as exception:
                    ws.cell(7, col + (rev_count * 3) + 1).value = "例外発生"
                if (comment_count > max_comment_count):
                    ws.row_dimensions[7].height = comment_count * const_row_heigt
                    max_comment_count = comment_count

                changed_functions = []
                for file_rslt in rslt.files:
                    if (file_rslt.file_name == file_with_module):
                        #/* 全体のステップ数（新規＋変更、削除、実ステップ）を出力 */
                        ws.cell(8, col + (rev_count * 3) + 1).value = file_rslt.file_steps.new_steps + file_rslt.file_steps.mod_steps
                        ws.cell(8, col + (rev_count * 3) + 2).value = file_rslt.file_steps.del_steps
                        ws.cell(8, col + (rev_count * 3) + 3).value = file_rslt.file_steps.real_steps

                        #/* モジュール（関数）単位のステップ数出力 */
                        for module_rslt in file_rslt.modules:
                            row = find_row(ws, col, 10, module_rslt.module_name)
                            ws.cell(row, col + (rev_count * 3) + 1).value = module_rslt.steps.new_steps + module_rslt.steps.mod_steps
                            ws.cell(row, col + (rev_count * 3) + 2).value = module_rslt.steps.del_steps
                            ws.cell(row, col + (rev_count * 3) + 3).value = module_rslt.steps.real_steps

                            #/* 最長モジュール名の長さに合わせてB列幅を調整 */
                            if (len(module_rslt.module_name) > max_file_name):
                                max_file_name = len(module_rslt.module_name)
                                ws.column_dimensions['B'].width = max_file_name * const_col_width

                            #/* 変更のあったモジュール（関数）をピックアップ(ついで末尾の節だけを拾って、型情報とかはそぎ落とす) */
                            if ((module_rslt.steps.new_steps + module_rslt.steps.mod_steps + module_rslt.steps.del_steps) > 0):
                                changed_function = module_rslt.module_name.split(" ")[-1]
                                changed_functions.append(changed_function)

                #/* セル結合して変更のあったモジュールの出力 */
                ws.merge_cells(start_row= 9, end_row=9, start_column=col + (rev_count * 3) + 1, end_column=col + (rev_count * 3) + 3)
                ws.cell(9, col + (rev_count * 3) + 1).alignment = Alignment(horizontal="left", vertical="top")
                changed = ""
                line_count = 1
                for changed_function in changed_functions:
                    if (changed == ""):
                        changed = changed_function
                    else:
                        changed += "\n" + changed_function
                        line_count += 1
                ws.cell(9, col + (rev_count * 3) + 1).value = changed

                if (line_count > max_line_count):
                    ws.row_dimensions[9].height = line_count * const_row_heigt
                    max_line_count = line_count

                rev_count += 1

        wb.save(g_out_xlsx_file)

    return


#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    global g_kazoe_only

    start_time = time.perf_counter()

    check_command_line_option()
    log_settings()

    if (g_kazoe_only == 0):
        check_repo_info()
        check_path_log()
        output_path_files()

    check_kazoe_result()
    out_kazoe_history()

    end_time = time.perf_counter()
    now = datetime.datetime.now()
    print("end checking : " + str(now))
    second = int(end_time - start_time)
    minute = second / 60
    second = second % 60
    print("  %dmin %dsec" % (minute, second))
    return


if __name__ == "__main__":
    main()
